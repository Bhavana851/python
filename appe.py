from pathlib import Path
path = Path()
for file in path.glob('*'):
    print(file)



import openpyxl as xl
wb = xl.load_workbook('transaction.xlsx')
sheet = wb['sheet1']
cell = sheet['a1']
cell = sheet.cell(1,1)
for row in range(1,sheet.max_row+1):
    print(row)
